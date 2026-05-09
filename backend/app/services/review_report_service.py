"""Generate downloadable readiness reports (PDF, XLSX) from a ProposalReview.

The LLM produces one combined Markdown blob (review_output) with sections
shaped like:

    ## 1. Criterion name
    Score: 7/10
    - **Status** — ✅ Pass
    - **Findings** — body
    - **Recommendations** — body

This module re-parses that blob into structured criteria (mirrors
frontend/src/utils/reviewOutput.ts) and renders two report formats:
  - PDF (reportlab): a presentable readiness report
  - XLSX (openpyxl): an analytical workbook with one row per criterion

Both formats are derivative of the same parsed structure.
"""
import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone

from app.models.proposal_review import ProposalReview


# ---------- Parser ----------

@dataclass
class ParsedCriterion:
    index: int
    name: str
    score: float | None
    status: str  # pass | partial | fail | na | unknown
    findings: str
    recommendations: str
    body: str


_HEADER_RE = re.compile(r"^##\s+(\d+)\.\s+(.+?)\s*$", re.MULTILINE)
_SCORE_RE = re.compile(r"[Ss]core:\s*([\d.]+)\s*/\s*10")


def parse_criteria(review_output: str | None) -> list[ParsedCriterion]:
    """Split combined review markdown into per-criterion records."""
    if not review_output:
        return []
    md = review_output

    matches = [
        (int(m.group(1)), m.group(2).strip(), m.start(), m.end())
        for m in _HEADER_RE.finditer(md)
    ]

    out: list[ParsedCriterion] = []
    for i, (idx, name, _start, head_end) in enumerate(matches):
        next_start = matches[i + 1][2] if i + 1 < len(matches) else len(md)
        body = md[head_end:next_start].strip()
        out.append(
            ParsedCriterion(
                index=idx,
                name=name,
                score=_extract_score(body),
                status=_extract_status(body),
                findings=_extract_section(body, "Findings"),
                recommendations=_extract_section(body, "Recommendations"),
                body=body,
            )
        )
    return out


def _extract_score(body: str) -> float | None:
    m = _SCORE_RE.search(body)
    if not m:
        return None
    try:
        v = float(m.group(1))
    except ValueError:
        return None
    if 0.0 <= v <= 10.0:
        return v
    return None


def _extract_status(body: str) -> str:
    head = body[:600]
    if re.search(r"Status[^\n]*(✅|\bPass\b)", head, re.IGNORECASE):
        return "pass"
    if re.search(r"Status[^\n]*(⚠|\bPartial\b)", head, re.IGNORECASE):
        return "partial"
    if re.search(r"Status[^\n]*(❌|\bFail\b)", head, re.IGNORECASE):
        return "fail"
    if re.search(r"Status[^\n]*(🟡|N/A|\bNA\b)", head, re.IGNORECASE):
        return "na"
    return "unknown"


def _extract_section(body: str, label: str) -> str:
    # Form 1: bulleted bold "- **Findings** — ..."
    pat1 = re.compile(
        rf"-\s*\*\*{re.escape(label)}\*\*\s*[—:\-]?\s*([\s\S]*?)"
        rf"(?=\n\s*-\s*\*\*[A-Z][a-zA-Z &]+\*\*|\n##\s|\Z)",
        re.IGNORECASE,
    )
    m = pat1.search(body)
    if m:
        return m.group(1).strip()
    # Form 2: heading-only "**Findings**" or "### Findings"
    pat2 = re.compile(
        rf"(?:\*\*{re.escape(label)}\*\*|###?\s*{re.escape(label)}\b)"
        rf"\s*[—:\-]?\s*([\s\S]*?)"
        rf"(?=\n\s*\*\*[A-Z][a-zA-Z &]+\*\*|\n###?\s|\n##\s|\Z)",
        re.IGNORECASE,
    )
    m = pat2.search(body)
    if m:
        return m.group(1).strip()
    return ""


def aggregate_score(criteria: list[ParsedCriterion]) -> float | None:
    scores = [c.score for c in criteria if c.score is not None]
    if not scores:
        return None
    return round(sum(scores) / len(scores), 1)


def readiness_buckets(criteria: list[ParsedCriterion]) -> dict[str, int]:
    what_if = moderate = good = 0
    for c in criteria:
        if c.score is None:
            continue
        if c.score < 5:
            what_if += 1
        elif c.score < 7:
            moderate += 1
        else:
            good += 1
    return {"what_if": what_if, "moderate": moderate, "good_to_pass": good}


def verdict(score: float | None) -> tuple[str, str]:
    """Return (label, tone). Tones: ready / go-edits / no-go / unknown."""
    if score is None:
        return ("PENDING", "unknown")
    if score >= 8.0:
        return ("READY TO SUBMIT", "ready")
    if score >= 7.0:
        return ("GO WITH EDITS", "go-edits")
    return ("NO GO DECISION", "no-go")


# ---------- Helpers shared by both renderers ----------

def _safe_filename_stem(review: ProposalReview) -> str:
    base = (review.source_filename or f"review-{review.id}").rsplit(".", 1)[0]
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", base).strip("_")
    return cleaned[:80] or f"review-{review.id}"


def _md(text: str) -> str:
    return text or "—"


# ---------- PDF renderer ----------

# KPMG palette (matches the SPA's tailwind config).
_KPMG_BLUE = "#00338D"
_KPMG_MEDIUM = "#005EB8"
_INK = "#1F1A14"
_MUTED = "#7F7666"
_LIGHT = "#F4F1EB"
_LINE = "#E2DDD3"
_READY = "#0E8A6B"
_GO_EDITS = "#B97700"
_NO_GO = "#A21515"

_VERDICT_COLOR = {
    "ready": _READY,
    "go-edits": _GO_EDITS,
    "no-go": _NO_GO,
    "unknown": _MUTED,
}


def _escape_rl(s: str) -> str:
    """Escape characters that ReportLab Paragraph treats as XML markup."""
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _md_to_rl(text: str) -> str:
    """Translate a tiny subset of Markdown to ReportLab inline markup."""
    s = _escape_rl(text)
    s = re.sub(r"\*\*([^*]+?)\*\*", r"<b>\1</b>", s)
    s = re.sub(r"(?<!\*)\*([^*\n]+?)\*(?!\*)", r"<i>\1</i>", s)
    s = re.sub(r"\n\s*[-*]\s+", "<br/>• ", s)
    s = s.replace("\n", "<br/>")
    return s


def build_pdf_report(review: ProposalReview) -> bytes:
    """Render the review as a multi-page PDF readiness report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    criteria = parse_criteria(review.review_output)
    overall = aggregate_score(criteria)
    buckets = readiness_buckets(criteria)
    verd_label, verd_tone = verdict(overall)
    verd_color_hex = _VERDICT_COLOR[verd_tone]

    KPMG_BLUE = colors.HexColor(_KPMG_BLUE)
    INK = colors.HexColor(_INK)
    MUTED = colors.HexColor(_MUTED)
    LIGHT = colors.HexColor(_LIGHT)
    LINE = colors.HexColor(_LINE)
    READY_COLOR = colors.HexColor(_READY)
    GO_EDITS_COLOR = colors.HexColor(_GO_EDITS)
    NO_GO_COLOR = colors.HexColor(_NO_GO)
    VERD = colors.HexColor(verd_color_hex)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=f"Proposal Readiness Report — {review.source_filename or review.id}",
    )

    base = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=base["Heading1"],
                        fontName="Helvetica-Bold", fontSize=22, leading=26,
                        textColor=INK, spaceAfter=2)
    h2 = ParagraphStyle("H2", parent=base["Heading2"],
                        fontName="Helvetica-Bold", fontSize=14, leading=18,
                        textColor=KPMG_BLUE, spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("Body", parent=base["BodyText"],
                          fontName="Helvetica", fontSize=10, leading=14,
                          textColor=INK, spaceAfter=4)
    muted = ParagraphStyle("Muted", parent=body, textColor=MUTED, fontSize=9)
    eyebrow = ParagraphStyle("Eyebrow", parent=body, fontName="Helvetica-Bold",
                             fontSize=8, textColor=MUTED, spaceAfter=2,
                             leading=10)
    score_big = ParagraphStyle("ScoreBig", parent=body,
                               fontName="Helvetica-Bold", fontSize=42, leading=46,
                               textColor=KPMG_BLUE, alignment=2)
    verd_style = ParagraphStyle("Verdict", parent=body,
                                fontName="Helvetica-Bold", fontSize=14, leading=18,
                                textColor=VERD)
    title_style = ParagraphStyle("Title", parent=h1,
                                 fontName="Helvetica-Bold", fontSize=15, leading=19,
                                 textColor=INK, spaceAfter=4)

    story: list = []

    # ---- Header ----
    story.append(Paragraph("Proposal Readiness Report", h1))
    story.append(Paragraph(
        f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}",
        muted,
    ))
    story.append(Spacer(1, 0.4 * cm))

    # ---- Document title + client ----
    md = review.extracted_metadata or {}
    title = md.get("document_title") or review.source_filename or "Untitled proposal"
    client = md.get("client_name") or "—"
    story.append(Paragraph(_escape_rl(title), title_style))
    story.append(Paragraph(f"<b>{_escape_rl(client)}</b>", body))
    story.append(Spacer(1, 0.3 * cm))

    # ---- Source file table ----
    src_data = [
        ["Source file", review.source_filename or "—"],
        ["Format", (review.source_kind or "").upper() or "—"],
        ["Size", f"{(review.source_size_bytes or 0) / (1024 * 1024):.2f} MB"],
        ["Model", review.model or "—"],
        ["Document class", (review.document_class or "—").title()],
        ["Reviewed at",
         review.created_at.strftime("%Y-%m-%d %H:%M") if review.created_at else "—"],
    ]
    src_tbl = Table(src_data, colWidths=[3.5 * cm, 13 * cm], hAlign="LEFT")
    src_tbl.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 9),
        ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
        ("TEXTCOLOR", (1, 0), (1, -1), INK),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, LINE),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, LINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(src_tbl)
    story.append(Spacer(1, 0.5 * cm))

    # ---- Verdict + score panel ----
    score_str = f"{overall:.1f}" if overall is not None else "—"
    summary_line = ""
    if overall is not None:
        if overall >= 8.0:
            summary_line = "Submission-ready. Light polish recommended."
        elif overall >= 7.0:
            summary_line = "Submission-ready, with edits. Address warnings before sending."
        else:
            summary_line = 'Critical "Must Fix" issues detected. Remediation is required prior to submission.'
    verd_panel_data = [
        [Paragraph("READINESS INDEX", eyebrow),
         Paragraph(f"{score_str}", score_big)],
        [Paragraph("VERDICT", eyebrow),
         Paragraph(verd_label, verd_style)],
    ]
    verd_panel = Table(verd_panel_data, colWidths=[10.5 * cm, 6 * cm])
    verd_panel.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(verd_panel)
    if summary_line:
        story.append(Paragraph(_escape_rl(summary_line), muted))
    story.append(Spacer(1, 0.4 * cm))

    # ---- Bucket distribution ----
    story.append(Paragraph("READINESS DISTRIBUTION", eyebrow))
    bucket_rows = [
        ["WHAT IF (<5)", "MODERATE (5–7)", "GOOD TO PASS (≥7)"],
        [str(buckets["what_if"]), str(buckets["moderate"]), str(buckets["good_to_pass"])],
    ]
    bucket_tbl = Table(bucket_rows, colWidths=[5.16 * cm, 5.16 * cm, 5.16 * cm])
    bucket_tbl.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 8),
        ("FONT", (0, 1), (-1, 1), "Helvetica-Bold", 22),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FBE5E5")),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#FBF1DC")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E6F2EC")),
        ("TEXTCOLOR", (0, 0), (0, -1), NO_GO_COLOR),
        ("TEXTCOLOR", (1, 0), (1, -1), GO_EDITS_COLOR),
        ("TEXTCOLOR", (2, 0), (2, -1), READY_COLOR),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING", (0, 1), (-1, 1), 0),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
    ]))
    story.append(bucket_tbl)
    story.append(Spacer(1, 0.5 * cm))

    # ---- Document metadata block (if any) ----
    meta_rows = []
    for label, key in [
        ("Submission date", "submission_date"),
        ("Purpose & scope", "purpose_and_scope"),
        ("Mandatory requirements", "client_mandatory_requirements"),
    ]:
        v = (md.get(key) or "").strip()
        if v:
            meta_rows.append((label, v))
    if meta_rows:
        story.append(Paragraph("Proposal metadata", h2))
        meta_tbl = Table(
            [[Paragraph(_escape_rl(lbl), eyebrow),
              Paragraph(_md_to_rl(val), body)] for lbl, val in meta_rows],
            colWidths=[3.5 * cm, 13 * cm],
            hAlign="LEFT",
        )
        meta_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, -2), 0.25, LINE),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 0.4 * cm))

    # ---- Criteria summary table ----
    if criteria:
        story.append(Paragraph("Criteria summary", h2))
        sum_rows: list = [["#", "Criterion", "Score", "Status"]]
        for c in criteria:
            score_cell = f"{c.score:.1f}/10" if c.score is not None else "—"
            status_cell = c.status.upper().replace("-", " ") if c.status != "unknown" else "—"
            sum_rows.append([
                str(c.index),
                Paragraph(_escape_rl(c.name), body),
                score_cell,
                status_cell,
            ])
        sum_tbl = Table(sum_rows, colWidths=[1 * cm, 11 * cm, 1.8 * cm, 2.7 * cm])
        sum_tbl.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
            ("BACKGROUND", (0, 0), (-1, 0), KPMG_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (3, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, LINE),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, LINE),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(sum_tbl)

        # ---- Per-criterion detail ----
        story.append(PageBreak())
        story.append(Paragraph("Detailed findings", h1))
        story.append(Spacer(1, 0.2 * cm))
        for i, c in enumerate(criteria):
            score_str = f"{c.score:.1f}/10" if c.score is not None else "—"
            score_color = (
                _READY if (c.score or 0) >= 7
                else _GO_EDITS if (c.score or 0) >= 5
                else _NO_GO if c.score is not None
                else _MUTED
            )
            story.append(Paragraph(
                f"{c.index}. {_escape_rl(c.name)}"
                f"  <font color='{score_color}' size='12'><b>{score_str}</b></font>",
                h2,
            ))
            if c.findings:
                story.append(Paragraph("FINDINGS", eyebrow))
                story.append(Paragraph(_md_to_rl(c.findings), body))
                story.append(Spacer(1, 0.15 * cm))
            if c.recommendations:
                story.append(Paragraph("RECOMMENDATIONS", eyebrow))
                story.append(Paragraph(_md_to_rl(c.recommendations), body))
                story.append(Spacer(1, 0.15 * cm))
            if not c.findings and not c.recommendations:
                # Fall back to the raw body so the user sees something.
                story.append(Paragraph(_md_to_rl(c.body), body))
            if i < len(criteria) - 1:
                story.append(Spacer(1, 0.4 * cm))
    else:
        story.append(Paragraph(
            "No structured criteria were parsed from this review. "
            "Open the review in the app to inspect the raw output.",
            muted,
        ))

    doc.build(story)
    return buf.getvalue()


# ---------- XLSX renderer ----------

def build_xlsx_report(review: ProposalReview) -> bytes:
    """Render the review as a 3-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    criteria = parse_criteria(review.review_output)
    overall = aggregate_score(criteria)
    buckets = readiness_buckets(criteria)
    verd_label, verd_tone = verdict(overall)
    md = review.extracted_metadata or {}

    bold = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="00338D")
    white_bold = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=16, color="00338D")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ---- Sheet 1: Summary ----
    s1 = wb.active
    s1.title = "Summary"

    rows: list[tuple] = [
        ("Proposal Readiness Report", ""),
        ("Generated", f"{datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}"),
        ("", ""),
        ("Document title", _md(md.get("document_title") or review.source_filename or "")),
        ("Client", _md(md.get("client_name") or "")),
        ("Submission date", _md(md.get("submission_date") or "")),
        ("Purpose & scope", _md(md.get("purpose_and_scope") or "")),
        ("Mandatory requirements", _md(md.get("client_mandatory_requirements") or "")),
        ("", ""),
        ("Source file", review.source_filename or "—"),
        ("Format", (review.source_kind or "").upper() or "—"),
        ("Size (MB)", round((review.source_size_bytes or 0) / (1024 * 1024), 2)),
        ("Model", review.model or "—"),
        ("Document class", review.document_class or "—"),
        ("Reviewed at",
         review.created_at.strftime("%Y-%m-%d %H:%M") if review.created_at else "—"),
        ("", ""),
        ("Overall score",
         f"{overall:.1f}/10" if overall is not None else "—"),
        ("Verdict", verd_label),
        ("WHAT IF (<5)", buckets["what_if"]),
        ("MODERATE (5–7)", buckets["moderate"]),
        ("GOOD TO PASS (≥7)", buckets["good_to_pass"]),
    ]
    for r in rows:
        s1.append(r)

    s1["A1"].font = title_font
    for ridx in range(1, len(rows) + 1):
        s1.cell(row=ridx, column=1).font = bold
    for ridx in range(4, 9):  # metadata block: wrap long values
        s1.cell(row=ridx, column=2).alignment = wrap_top
    s1.column_dimensions["A"].width = 28
    s1.column_dimensions["B"].width = 70

    # ---- Sheet 2: Criteria ----
    s2 = wb.create_sheet("Criteria")
    headers = ["#", "Criterion", "Score", "Status", "Findings", "Recommendations"]
    s2.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = s2.cell(row=1, column=col_idx)
        c.font = white_bold
        c.fill = blue_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for cri in criteria:
        s2.append([
            cri.index,
            cri.name,
            f"{cri.score:.1f}/10" if cri.score is not None else "—",
            cri.status.upper().replace("-", " ") if cri.status != "unknown" else "—",
            cri.findings or "",
            cri.recommendations or "",
        ])

    for col, w in zip("ABCDEF", [4, 32, 9, 12, 60, 60]):
        s2.column_dimensions[col].width = w
    for row in s2.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.alignment = wrap_top
    s2.freeze_panes = "A2"

    # ---- Sheet 3: Full review markdown ----
    s3 = wb.create_sheet("Full review (markdown)")
    s3["A1"] = "Full LLM output"
    s3["A1"].font = bold
    s3["A2"] = review.review_output or ""
    s3["A2"].alignment = wrap_top
    s3.column_dimensions["A"].width = 120
    s3.row_dimensions[2].height = 600

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def report_filename(review: ProposalReview, ext: str) -> str:
    """Stable filename for downloads, e.g. AI_Strategy_Proposal_review.pdf."""
    return f"{_safe_filename_stem(review)}_review.{ext}"
