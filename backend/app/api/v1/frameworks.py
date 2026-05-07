"""Review framework endpoints: CRUD + AI-assisted criteria generation + export/import."""
import io
import json
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, File, HTTPException, UploadFile, status
from fastapi.responses import Response, StreamingResponse

from app.core.deps import CurrentUser, DbSession
from app.schemas.framework import (
    AutoGenFromFileResponse,
    FrameworkCreate,
    FrameworkCriterion,
    FrameworkListResponse,
    FrameworkResponse,
    FrameworkSummary,
    FrameworkUpdate,
)
from app.services import file_parser_service, framework_service, ollama_service

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/frameworks", tags=["frameworks"])


def _summary(row) -> FrameworkSummary:
    return FrameworkSummary(
        id=row.id,
        owner_user_id=row.owner_user_id,
        name=row.name,
        is_public=row.is_public,
        criteria_count=len(row.criteria or []),
        updated_at=row.updated_at,
    )


@router.get("", response_model=FrameworkListResponse)
async def list_frameworks(
    db: DbSession,
    user: CurrentUser,
    search: str | None = None,
    limit: int = 100,
    offset: int = 0,
):
    rows, total = await framework_service.list_visible(
        db, user=user, search=search, limit=min(limit, 500), offset=max(offset, 0)
    )
    return FrameworkListResponse(items=[_summary(r) for r in rows], total=total)


@router.post("", response_model=FrameworkResponse, status_code=status.HTTP_201_CREATED)
async def create_framework(req: FrameworkCreate, db: DbSession, user: CurrentUser):
    return await framework_service.create(db, user=user, req=req)


@router.get("/{framework_id}", response_model=FrameworkResponse)
async def get_framework(framework_id: int, db: DbSession, user: CurrentUser):
    item = await framework_service.get_visible(db, user=user, framework_id=framework_id)
    if not item:
        raise HTTPException(status_code=404, detail="Framework not found.")
    return item


@router.patch("/{framework_id}", response_model=FrameworkResponse)
async def update_framework(
    framework_id: int, req: FrameworkUpdate, db: DbSession, user: CurrentUser
):
    item = await framework_service.get_visible(db, user=user, framework_id=framework_id)
    if not item:
        raise HTTPException(status_code=404, detail="Framework not found.")
    if item.owner_user_id and item.owner_user_id != user.id and not user.is_superadmin:
        raise HTTPException(
            status_code=403,
            detail="You can only edit frameworks you own (admins can edit any).",
        )
    return await framework_service.update(db, item=item, req=req)


@router.delete("/{framework_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_framework(framework_id: int, db: DbSession, user: CurrentUser):
    item = await framework_service.get_visible(db, user=user, framework_id=framework_id)
    if not item:
        raise HTTPException(status_code=404, detail="Framework not found.")
    if item.owner_user_id and item.owner_user_id != user.id and not user.is_superadmin:
        raise HTTPException(
            status_code=403,
            detail="You can only delete frameworks you own (admins can delete any).",
        )
    await framework_service.delete(db, item=item)


@router.post(
    "/auto-gen",
    response_model=AutoGenFromFileResponse,
    summary="Generate criteria by analysing a sample document with the LLM",
)
async def auto_gen_criteria(
    user: CurrentUser,
    file: UploadFile = File(..., description="Sample proposal/checklist (.pptx/.docx/.pdf)"),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename.")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    try:
        criteria = await framework_service.autogen_from_file(
            filename=file.filename, file_bytes=raw
        )
    except file_parser_service.UnsupportedFile as e:
        raise HTTPException(status_code=415, detail=str(e))
    except file_parser_service.FileTooLarge as e:
        raise HTTPException(status_code=413, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except ollama_service.OllamaError as e:
        logger.exception("Ollama failure during auto-gen")
        raise HTTPException(
            status_code=502, detail=f"AI service unavailable: {e}"
        )
    return AutoGenFromFileResponse(criteria=criteria)


# -------- Export / Import --------

@router.get(
    "/{framework_id}/export",
    summary="Export a framework as JSON or Excel",
)
async def export_framework(
    framework_id: int,
    db: DbSession,
    user: CurrentUser,
    format: str = "json",
):
    item = await framework_service.get_visible(db, user=user, framework_id=framework_id)
    if not item:
        raise HTTPException(status_code=404, detail="Framework not found.")

    # Parse criteria through the Pydantic model for consistent shape
    criteria = [
        FrameworkCriterion.model_validate(c).model_dump()
        for c in (item.criteria or [])
    ]

    safe_name = item.name.replace(" ", "_")[:40]

    if format == "xlsx":
        return _export_xlsx(item, criteria, safe_name)

    # Default: JSON
    payload = {
        "schema_version": "2.0",
        "name": item.name,
        "persona_instruction": item.persona_instruction,
        "model": item.model,
        "is_public": item.is_public,
        "criteria": criteria,
        "exported_at": datetime.now(timezone.utc).isoformat(),
    }
    content = json.dumps(payload, ensure_ascii=False, indent=2)
    return Response(
        content=content,
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}.json"',
        },
    )


def _export_xlsx(item, criteria: list[dict], safe_name: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()

    # Sheet 1: Framework metadata
    ws_meta = wb.active
    ws_meta.title = "Framework"
    ws_meta.append(["Field", "Value"])
    ws_meta["A1"].font = Font(bold=True)
    ws_meta["B1"].font = Font(bold=True)
    ws_meta.append(["Name", item.name])
    ws_meta.append(["Persona Instruction", item.persona_instruction])
    ws_meta.append(["Model", item.model])
    ws_meta.append(["Public", str(item.is_public)])
    ws_meta.column_dimensions["A"].width = 25
    ws_meta.column_dimensions["B"].width = 80

    # Sheet 2: Criteria
    ws_crit = wb.create_sheet("Criteria")
    headers = [
        "Group", "Name (EN)", "Name (AR)",
        "Description (EN)", "Description (AR)",
        "Prompt (EN)", "Prompt (AR)",
    ]
    ws_crit.append(headers)
    for i, h in enumerate(headers, 1):
        ws_crit.cell(row=1, column=i).font = Font(bold=True)

    for c in criteria:
        ws_crit.append([
            c.get("group", ""),
            c.get("name_en", ""),
            c.get("name_ar", ""),
            c.get("description_en", ""),
            c.get("description_ar", ""),
            c.get("prompt_instruction_en", ""),
            c.get("prompt_instruction_ar", ""),
        ])

    # Set Arabic columns to RTL alignment
    for row in ws_crit.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    for row in ws_crit.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    for row in ws_crit.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    # Auto-size columns
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_crit.column_dimensions[col_letter].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}.xlsx"',
        },
    )


@router.post(
    "/import",
    response_model=FrameworkResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Import a framework from a JSON file",
)
async def import_framework(
    db: DbSession,
    user: CurrentUser,
    file: UploadFile = File(..., description="Framework JSON file"),
):
    if not file.filename or not file.filename.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="Only .json files are accepted.")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="Invalid JSON file.")

    if not isinstance(data, dict):
        raise HTTPException(status_code=422, detail="Expected a JSON object.")

    name = data.get("name")
    if not name or not isinstance(name, str):
        raise HTTPException(status_code=422, detail="Missing or invalid 'name' field.")

    # Parse criteria through the Pydantic model for validation + legacy migration
    raw_criteria = data.get("criteria", [])
    if not isinstance(raw_criteria, list):
        raise HTTPException(status_code=422, detail="'criteria' must be an array.")

    criteria: list[FrameworkCriterion] = []
    for i, c in enumerate(raw_criteria):
        try:
            criteria.append(FrameworkCriterion.model_validate(c))
        except Exception as e:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid criterion at index {i}: {e}",
            )

    req = FrameworkCreate(
        name=name.strip(),
        persona_instruction=data.get("persona_instruction", ""),
        model=data.get("model", "gemma4:latest"),
        is_public=False,  # Imported frameworks are always private initially
        criteria=criteria,
    )
    return await framework_service.create(db, user=user, req=req)
